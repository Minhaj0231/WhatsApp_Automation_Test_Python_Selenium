import openpyxl

class Excel(object):
    def __init__(self, path):
        self.path = path
        self.exel_file =  openpyxl.load_workbook(path)
        self.sheet = self.exel_file.get_sheet_by_name("Sheet1")    


    def get_mobile_number(self,row, col):

        number = self.sheet.cell(row=row, column=col).value
        return number      

    def write_in_excel_sheet(self, message, row, col): 

        self.sheet.cell(row=row, column=col).value = message
        self.exel_file.save(self.path)
        return True 